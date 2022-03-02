import os
import openpyxl

path = r"E:\studyDemo"
os.chdir(path)  # 修改工作路径
fileName = 'test.xlsx'
sheetName = 'Sheet1'


workbook = openpyxl.load_workbook(fileName)  # 返回一个workbook数据类型的值
worksheet = workbook[sheetName]  # 获取活动表
copyWorkSheet = workbook.copy_worksheet(worksheet)
print(copyWorkSheet)


rows = copyWorkSheet.max_row
col = copyWorkSheet.max_row


# 单元格拆分
emptyDataList = copyWorkSheet.merged_cells
print(emptyDataList)
if emptyDataList:
    emData = []
    for list in emptyDataList:
        r1, r2, c1, c2 = list.min_row, list.max_row, list.min_col, list.max_col
        emData.append((r1, r2, c1, c2))
        print(r1, r2, c1, c2)

    for r in emData:
        copyWorkSheet.unmerge_cells(
            start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

        for i in range(r[1] - r[0] + 1):
            for j in range(r[3] - r[2] + 1):
                copyWorkSheet.cell(row=r[0] + i, column=r[2] + j,
                                   value=copyWorkSheet.cell(r[0], r[2]).value)


workbook.save(fileName)
